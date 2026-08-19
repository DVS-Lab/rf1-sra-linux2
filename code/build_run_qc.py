#!/usr/bin/env python3
"""Build and verify canonical cohort-level RF1-SRA run imaging QC."""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.metadata
import json
import math
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


TARGET_SPACE = "MNI152NLin6Asym"
TARGET_MASK_NAME = "rf1-sra_MNI152NLin6Asym_desc-qctarget_mask.nii.gz"
METRIC_FLAGS = {
    "tsnr": "tsnr_outlier",
    "fd_mean": "fd_mean_outlier",
    "tedana_rejected_components": "tedana_outlier",
    "brain_coverage_pct": "brain_coverage_outlier",
}
RUN_COLUMNS = [
    "subject",
    "session",
    "paradigm",
    "task",
    "run",
    "tsnr",
    "fd_mean",
    "tedana_total_components",
    "tedana_accepted_components",
    "tedana_rejected_components",
    "tedana_rejected_fraction",
    "brain_coverage_pct",
    "tsnr_outlier",
    "fd_mean_outlier",
    "tedana_outlier",
    "brain_coverage_outlier",
    "imaging_qc_outlier",
    "outlier_reasons",
    "qc_complete",
    "missing_metrics",
    "qc_status",
    "bids_bold",
    "mriqc_json",
    "tedana_metrics",
    "fmriprep_brain_mask",
]
THRESHOLD_COLUMNS = [
    "paradigm",
    "bids_tasks",
    "metric",
    "n",
    "q1",
    "q3",
    "iqr",
    "lower_fence",
    "upper_fence",
    "outlier_direction",
    "n_outliers",
]
PAIR_COLUMNS = [
    "subject",
    "session",
    "socialdoors_present",
    "doors_present",
    "socialdoors_imaging_qc_outlier",
    "doors_imaging_qc_outlier",
    "socialdoors_outlier_reasons",
    "doors_outlier_reasons",
    "either_run_imaging_qc_outlier",
    "both_runs_imaging_qc_pass",
    "pair_qc_complete",
    "pair_issue",
    "socialdoors_tsnr",
    "doors_tsnr",
    "socialdoors_fd_mean",
    "doors_fd_mean",
    "socialdoors_tedana_rejected_components",
    "doors_tedana_rejected_components",
    "socialdoors_brain_coverage_pct",
    "doors_brain_coverage_pct",
]
CANONICAL_OUTPUTS = [
    Path("run_qc.tsv"),
    Path("thresholds.tsv"),
    Path("socialdoors_pair_qc.tsv"),
    Path("provenance.json"),
    Path("reference") / TARGET_MASK_NAME,
    *[
        Path("spreadsheets") / f"{name}_qc.xlsx"
        for name in ("sharedreward", "trust", "ugr", "socialdoors")
    ],
    *[
        Path("figures") / f"{name}_histograms.png"
        for name in ("sharedreward", "trust", "ugr", "socialdoors")
    ],
]


@dataclass(frozen=True, order=True)
class RunKey:
    subject: str
    session: str
    task: str
    run: str

    @property
    def prefix(self) -> str:
        return f"sub-{self.subject}_ses-{self.session}_task-{self.task}_run-{self.run}"


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_entities(name: str) -> dict[str, str]:
    stem = name
    for suffix in (".nii.gz", ".json", ".tsv"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
            break
    entities: dict[str, str] = {}
    for token in stem.split("_"):
        if "-" in token:
            key, value = token.split("-", 1)
            entities[key] = value
    return entities


def key_from_path(path: Path) -> RunKey | None:
    entities = parse_entities(path.name)
    required = ("sub", "ses", "task", "run")
    if not all(entities.get(name) for name in required):
        return None
    return RunKey(entities["sub"], entities["ses"], entities["task"], entities["run"])


def load_policy(path: Path) -> dict[str, Any]:
    policy = json.loads(path.read_text())
    if policy.get("schema_version") != 1:
        raise ValueError(
            f"unsupported QC policy schema: {policy.get('schema_version')}"
        )
    if policy.get("iqr_multiplier") != 1.5:
        raise ValueError("production QC policy requires an IQR multiplier of 1.5")
    if policy.get("quartile_method") != "linear" or policy.get("threshold_passes") != 1:
        raise ValueError("production QC policy requires linear quartiles in one pass")
    paradigms = policy.get("paradigms", {})
    expected_paradigms = {
        "sharedreward": ["sharedreward"],
        "trust": ["trust"],
        "ugr": ["ugr"],
        "socialdoors": ["socialdoors", "doors"],
    }
    if paradigms != expected_paradigms:
        raise ValueError("QC policy paradigm/task mapping is not canonical")
    task_map: dict[str, str] = {}
    for paradigm, tasks in paradigms.items():
        for task in tasks:
            if task in task_map:
                raise ValueError(f"task appears in multiple paradigms: {task}")
            task_map[task] = paradigm
    metrics = policy.get("metrics", {})
    if set(metrics) != set(METRIC_FLAGS):
        raise ValueError("QC policy metrics do not match the production metric set")
    expected_directions = {
        "tsnr": "lower",
        "fd_mean": "upper",
        "tedana_rejected_components": "upper",
        "brain_coverage_pct": "lower",
    }
    for metric, direction in expected_directions.items():
        if metrics[metric].get("direction") != direction:
            raise ValueError(f"noncanonical outlier direction for {metric}")
    policy["task_map"] = task_map
    return policy


def relative_path(path: Path | None, project_root: Path) -> str:
    if path is None:
        return ""
    try:
        return path.resolve().relative_to(project_root.resolve()).as_posix()
    except ValueError:
        return path.name


def source_excluded_subjects(root: Path) -> set[str]:
    if not root.is_dir():
        return set()
    subjects: set[str] = set()
    for path in root.glob("Smith-SRA-*"):
        if not path.is_dir():
            continue
        match = re.match(r"Smith-SRA-([0-9]+)(?:-|$)", path.name)
        if match:
            subjects.add(match.group(1))
    return subjects


def index_run_files(
    root: Path, pattern: str, tasks: set[str], predicate: Any | None = None
) -> dict[RunKey, list[Path]]:
    index: dict[RunKey, list[Path]] = {}
    if not root.is_dir():
        return index
    for path in sorted(root.rglob(pattern)):
        if not path.is_file():
            continue
        key = key_from_path(path)
        if key is None or key.task not in tasks:
            continue
        entities = parse_entities(path.name)
        if predicate is not None and not predicate(path, entities):
            continue
        index.setdefault(key, []).append(path)
    return index


def inventory_bids_runs(
    bids_root: Path,
    tasks: set[str],
    excluded_root: Path,
    include_source_excluded: bool = False,
) -> tuple[dict[RunKey, list[Path]], set[str]]:
    inventory = index_run_files(
        bids_root,
        "*_bold.nii.gz",
        tasks,
        lambda _path, ent: ent.get("echo") == "2" and ent.get("part") == "mag",
    )
    excluded = source_excluded_subjects(excluded_root)
    if not include_source_excluded:
        inventory = {
            key: paths
            for key, paths in inventory.items()
            if key.subject not in excluded
        }
    return inventory, excluded


def finite_float(value: Any) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def extract_mriqc(paths: list[Path]) -> tuple[dict[str, float | None], list[str]]:
    values = {"tsnr": None, "fd_mean": None}
    if len(paths) != 1:
        issue = "missing_mriqc_json" if not paths else "ambiguous_mriqc_json"
        return values, [f"tsnr:{issue}", f"fd_mean:{issue}"]
    try:
        data = json.loads(paths[0].read_text())
    except Exception:
        return values, ["tsnr:invalid_mriqc_json", "fd_mean:invalid_mriqc_json"]
    missing: list[str] = []
    for metric in values:
        values[metric] = finite_float(data.get(metric))
        if values[metric] is None:
            missing.append(f"{metric}:missing_or_nonfinite")
    return values, missing


def extract_tedana(
    paths: list[Path],
) -> tuple[dict[str, float | int | None], list[str]]:
    empty = {
        "tedana_total_components": None,
        "tedana_accepted_components": None,
        "tedana_rejected_components": None,
        "tedana_rejected_fraction": None,
    }
    if len(paths) != 1:
        issue = "missing_tedana_metrics" if not paths else "ambiguous_tedana_metrics"
        return empty, [f"tedana_rejected_components:{issue}"]
    try:
        with paths[0].open(newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            if reader.fieldnames is None or "classification" not in reader.fieldnames:
                raise ValueError("classification column missing")
            classifications = [
                str(row.get("classification", "")).strip().lower() for row in reader
            ]
    except Exception:
        return empty, ["tedana_rejected_components:invalid_classification_schema"]
    if not classifications or any(
        value not in {"accepted", "rejected"} for value in classifications
    ):
        return empty, ["tedana_rejected_components:invalid_classification_schema"]
    accepted = classifications.count("accepted")
    rejected = classifications.count("rejected")
    total = len(classifications)
    return {
        "tedana_total_components": total,
        "tedana_accepted_components": accepted,
        "tedana_rejected_components": rejected,
        "tedana_rejected_fraction": rejected / total,
    }, []


def science_modules() -> tuple[Any, Any, Any, Any, Any]:
    try:
        import nibabel as nib
        import numpy as np
        import pandas as pd
        from matplotlib import pyplot as plt
        from nibabel.processing import resample_from_to
    except ImportError as exc:
        raise RuntimeError(
            "build_run_qc.py needs numpy, pandas, nibabel, scipy, and matplotlib; "
            "use the shared TEDANA environment"
        ) from exc
    return np, pd, nib, resample_from_to, plt


def save_binary_image(data: Any, reference: Any, path: Path, nib: Any, np: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    image = nib.Nifti1Image(np.asarray(data, dtype=np.uint8), reference.affine)
    qform, qcode = reference.get_qform(coded=True)
    sform, scode = reference.get_sform(coded=True)
    image.set_qform(qform, int(qcode or 1))
    image.set_sform(sform, int(scode or 1))
    nib.save(image, str(path))


def build_target_mask(
    template_path: Path, exclusion_path: Path, output_path: Path
) -> dict[str, Any]:
    np, _pd, nib, resample_from_to, _plt = science_modules()
    template = nib.load(str(template_path))
    exclusion = nib.load(str(exclusion_path))
    if len(template.shape) != 3 or len(exclusion.shape) != 3:
        raise ValueError("coverage template and exclusion masks must be 3D")
    template_data = np.asanyarray(template.dataobj) > 0
    if template.shape != exclusion.shape or not np.allclose(
        template.affine, exclusion.affine, atol=1e-5
    ):
        exclusion = resample_from_to(exclusion, template, order=0)
    exclusion_data = np.asanyarray(exclusion.dataobj) > 0
    target = template_data & ~exclusion_data
    template_voxels = int(np.count_nonzero(template_data))
    excluded_voxels = int(np.count_nonzero(template_data & exclusion_data))
    target_voxels = int(np.count_nonzero(target))
    if not target_voxels or not excluded_voxels or target_voxels >= template_voxels:
        raise ValueError(
            "coverage target recipe did not remove a nonempty in-brain region"
        )
    save_binary_image(target, template, output_path, nib, np)
    return {
        "template_brain_mask_name": template_path.name,
        "template_brain_mask_sha256": sha256_file(template_path),
        "exclusion_mask_name": exclusion_path.name,
        "exclusion_mask_sha256": sha256_file(exclusion_path),
        "template_voxels": template_voxels,
        "excluded_in_brain_voxels": excluded_voxels,
        "target_voxels": target_voxels,
        "target_mask_sha256": sha256_file(output_path),
    }


def compute_coverage(target_path: Path, run_mask_path: Path) -> float:
    np, _pd, nib, resample_from_to, _plt = science_modules()
    target = nib.load(str(target_path))
    run_mask = nib.load(str(run_mask_path))
    if len(run_mask.shape) != 3:
        raise ValueError("run brain mask is not 3D")
    if target.shape != run_mask.shape or not np.allclose(
        target.affine, run_mask.affine, atol=1e-5
    ):
        target = resample_from_to(target, run_mask, order=0)
    target_data = np.asanyarray(target.dataobj) > 0
    run_data = np.asanyarray(run_mask.dataobj) > 0
    denominator = int(np.count_nonzero(target_data))
    if denominator == 0:
        raise ValueError("resampled coverage target is empty")
    covered = int(np.count_nonzero(target_data & run_data))
    coverage = 100.0 * covered / denominator
    if not 0.0 <= coverage <= 100.0:
        raise ValueError(f"impossible brain coverage: {coverage}")
    return coverage


def linear_quantile(values: Iterable[float], q: float) -> float:
    ordered = sorted(float(value) for value in values if math.isfinite(float(value)))
    if not ordered:
        raise ValueError("cannot calculate a quantile from no finite values")
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1.0 - weight) + ordered[upper] * weight


def compute_thresholds(
    rows: list[dict[str, Any]], policy: dict[str, Any]
) -> list[dict[str, Any]]:
    multiplier = float(policy["iqr_multiplier"])
    thresholds: list[dict[str, Any]] = []
    for paradigm, tasks in policy["paradigms"].items():
        group = [row for row in rows if row["paradigm"] == paradigm]
        for metric, spec in policy["metrics"].items():
            values = [
                float(row[metric])
                for row in group
                if row.get(metric) is not None and math.isfinite(float(row[metric]))
            ]
            if values:
                q1 = linear_quantile(values, 0.25)
                q3 = linear_quantile(values, 0.75)
                iqr = q3 - q1
                lower = q1 - multiplier * iqr
                upper = q3 + multiplier * iqr
            else:
                q1 = q3 = iqr = lower = upper = None
            thresholds.append(
                {
                    "paradigm": paradigm,
                    "bids_tasks": ";".join(tasks),
                    "metric": metric,
                    "n": len(values),
                    "q1": q1,
                    "q3": q3,
                    "iqr": iqr,
                    "lower_fence": lower,
                    "upper_fence": upper,
                    "outlier_direction": spec["direction"],
                    "n_outliers": 0,
                }
            )
    return thresholds


def apply_thresholds(
    rows: list[dict[str, Any]], thresholds: list[dict[str, Any]], policy: dict[str, Any]
) -> None:
    lookup = {(row["paradigm"], row["metric"]): row for row in thresholds}
    for row in rows:
        reasons: list[str] = []
        for metric, spec in policy["metrics"].items():
            value = row.get(metric)
            threshold = lookup[(row["paradigm"], metric)]
            flag: bool | None = None
            if value is not None and threshold["n"]:
                if spec["direction"] == "lower":
                    flag = float(value) < float(threshold["lower_fence"])
                elif spec["direction"] == "upper":
                    flag = float(value) > float(threshold["upper_fence"])
                else:
                    raise ValueError(f"unknown outlier direction: {spec['direction']}")
            row[METRIC_FLAGS[metric]] = flag
            if flag:
                reasons.append(spec["reason"])
                threshold["n_outliers"] += 1
        row["imaging_qc_outlier"] = any(
            row.get(flag) is True for flag in METRIC_FLAGS.values()
        )
        row["outlier_reasons"] = ";".join(reasons)
        row["qc_complete"] = not bool(row["missing_metrics"])
        row["qc_status"] = (
            "incomplete"
            if not row["qc_complete"]
            else "outlier" if row["imaging_qc_outlier"] else "pass"
        )


def build_socialdoors_pairs(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, list[dict[str, Any]]]] = {}
    for row in rows:
        if row["task"] not in {"socialdoors", "doors"}:
            continue
        pair = grouped.setdefault(
            (row["subject"], row["session"]), {"socialdoors": [], "doors": []}
        )
        pair[row["task"]].append(row)
    result: list[dict[str, Any]] = []
    metrics = ("tsnr", "fd_mean", "tedana_rejected_components", "brain_coverage_pct")
    for (subject, session), pair in sorted(grouped.items()):
        row: dict[str, Any] = {"subject": subject, "session": session}
        issues: list[str] = []
        selected: dict[str, dict[str, Any] | None] = {}
        for task in ("socialdoors", "doors"):
            entries = pair[task]
            row[f"{task}_present"] = bool(entries)
            selected[task] = entries[0] if len(entries) == 1 else None
            if not entries:
                issues.append(f"{task}_missing")
            elif len(entries) > 1:
                issues.append(f"{task}_multiple_runs")
            item = selected[task]
            if item is not None and not item["qc_complete"]:
                issues.append(f"{task}_qc_incomplete")
            row[f"{task}_imaging_qc_outlier"] = (
                item["imaging_qc_outlier"] if item is not None else None
            )
            row[f"{task}_outlier_reasons"] = (
                item["outlier_reasons"] if item is not None else ""
            )
            for metric in metrics:
                row[f"{task}_{metric}"] = item.get(metric) if item is not None else None
        known = [
            row["socialdoors_imaging_qc_outlier"],
            row["doors_imaging_qc_outlier"],
        ]
        row["either_run_imaging_qc_outlier"] = any(value is True for value in known)
        row["pair_qc_complete"] = not issues and all(
            selected[task] is not None and selected[task]["qc_complete"]
            for task in ("socialdoors", "doors")
        )
        row["both_runs_imaging_qc_pass"] = bool(
            row["pair_qc_complete"]
            and not row["socialdoors_imaging_qc_outlier"]
            and not row["doors_imaging_qc_outlier"]
        )
        row["pair_issue"] = ";".join(issues)
        result.append(row)
    return result


def build_rows(
    project_root: Path,
    policy: dict[str, Any],
    target_mask: Path,
    excluded_root: Path,
    include_source_excluded: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    tasks = set(policy["task_map"])
    bids_root = project_root / "bids"
    mriqc_root = project_root / "derivatives" / "mriqc"
    tedana_root = project_root / "derivatives" / "tedana"
    fmriprep_root = project_root / "derivatives" / "fmriprep"
    inventory, excluded = inventory_bids_runs(
        bids_root, tasks, excluded_root, include_source_excluded
    )
    mriqc = index_run_files(
        mriqc_root,
        "*_bold.json",
        tasks,
        lambda _path, ent: ent.get("echo") == "2" and ent.get("part") == "mag",
    )
    tedana = index_run_files(tedana_root, "*_desc-tedana_metrics.tsv", tasks)
    masks = index_run_files(
        fmriprep_root,
        "*_desc-brain_mask.nii.gz",
        tasks,
        lambda _path, ent: ent.get("space") == TARGET_SPACE and "echo" not in ent,
    )
    rows: list[dict[str, Any]] = []
    for key, bold_paths in sorted(inventory.items()):
        missing: list[str] = []
        if len(bold_paths) != 1:
            missing.append("bids_bold:ambiguous_inventory")
        mriqc_values, mriqc_missing = extract_mriqc(mriqc.get(key, []))
        tedana_values, tedana_missing = extract_tedana(tedana.get(key, []))
        missing.extend(mriqc_missing)
        missing.extend(tedana_missing)
        mask_paths = masks.get(key, [])
        coverage: float | None = None
        if len(mask_paths) != 1:
            issue = (
                "missing_fmriprep_brain_mask"
                if not mask_paths
                else "ambiguous_fmriprep_brain_mask"
            )
            missing.append(f"brain_coverage_pct:{issue}")
        else:
            try:
                coverage = compute_coverage(target_mask, mask_paths[0])
            except Exception as exc:
                missing.append(f"brain_coverage_pct:invalid_mask:{type(exc).__name__}")
        row: dict[str, Any] = {
            "subject": key.subject,
            "session": key.session,
            "paradigm": policy["task_map"][key.task],
            "task": key.task,
            "run": key.run,
            **mriqc_values,
            **tedana_values,
            "brain_coverage_pct": coverage,
            "missing_metrics": ";".join(sorted(set(missing))),
            "bids_bold": ";".join(
                relative_path(path, project_root) for path in bold_paths
            ),
            "mriqc_json": ";".join(
                relative_path(path, project_root) for path in mriqc.get(key, [])
            ),
            "tedana_metrics": ";".join(
                relative_path(path, project_root) for path in tedana.get(key, [])
            ),
            "fmriprep_brain_mask": ";".join(
                relative_path(path, project_root) for path in mask_paths
            ),
        }
        rows.append(row)
    return rows, {
        "inventory_runs": len(rows),
        "source_excluded_subjects": sorted(excluded),
        "source_excluded_runs_omitted": sum(
            1
            for key in index_run_files(
                bids_root,
                "*_bold.nii.gz",
                tasks,
                lambda _path, ent: ent.get("echo") == "2" and ent.get("part") == "mag",
            )
            if key.subject in excluded and not include_source_excluded
        ),
    }


def output_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return f"{value:.10g}"
    return value


def write_tsv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=columns, delimiter="\t", lineterminator="\n"
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {column: output_value(row.get(column)) for column in columns}
            )


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def normalize_xlsx(path: Path) -> None:
    temp = path.with_name(f".{path.name}.normalized")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as destination:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            content = source.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>1980-01-01T00:00:00Z\g<2>",
                    content,
                )
            destination.writestr(info, content)
    os.replace(temp, path)


def style_workbook(path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX output requires openpyxl in the QC Python environment"
        ) from exc
    workbook = load_workbook(path)
    fixed_time = datetime(1980, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    outlier_fill = PatternFill("solid", fgColor="F4CCCC")
    incomplete_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in range(1, sheet.max_column + 1):
            values = [
                str(sheet.cell(row, column).value or "")
                for row in range(1, sheet.max_row + 1)
            ]
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(10, max(map(len, values)) + 2), 45
            )
        headers = {cell.value: cell.column for cell in sheet[1]}
        if sheet.max_row > 1 and "imaging_qc_outlier" in headers:
            column = get_column_letter(headers["imaging_qc_outlier"])
            sheet.conditional_formatting.add(
                f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}",
                FormulaRule(formula=[f"${column}2=TRUE"], fill=outlier_fill),
            )
        if sheet.max_row > 1 and "qc_status" in headers:
            column = get_column_letter(headers["qc_status"])
            sheet.conditional_formatting.add(
                f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}",
                FormulaRule(formula=[f'${column}2="incomplete"'], fill=incomplete_fill),
            )
    workbook.save(path)
    normalize_xlsx(path)


def write_workbooks(
    directory: Path,
    rows: list[dict[str, Any]],
    thresholds: list[dict[str, Any]],
    pairs: list[dict[str, Any]],
) -> None:
    _np, pd, _nib, _resample, _plt = science_modules()
    directory.mkdir(parents=True, exist_ok=True)
    for paradigm in ("sharedreward", "trust", "ugr", "socialdoors"):
        path = directory / f"{paradigm}_qc.xlsx"
        run_rows = [row for row in rows if row["paradigm"] == paradigm]
        threshold_rows = [row for row in thresholds if row["paradigm"] == paradigm]
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame(run_rows, columns=RUN_COLUMNS).to_excel(
                writer, sheet_name="runs", index=False
            )
            if paradigm == "socialdoors":
                pd.DataFrame(pairs, columns=PAIR_COLUMNS).to_excel(
                    writer, sheet_name="paired_summary", index=False
                )
            pd.DataFrame(threshold_rows, columns=THRESHOLD_COLUMNS).to_excel(
                writer, sheet_name="thresholds", index=False
            )
        style_workbook(path)


def write_histograms(
    directory: Path,
    rows: list[dict[str, Any]],
    thresholds: list[dict[str, Any]],
) -> None:
    np, _pd, _nib, _resample, plt = science_modules()
    directory.mkdir(parents=True, exist_ok=True)
    labels = {
        "tsnr": "tSNR",
        "fd_mean": "Mean framewise displacement",
        "tedana_rejected_components": "TEDANA rejected components",
        "brain_coverage_pct": "Brain coverage (%)",
    }
    lookup = {(row["paradigm"], row["metric"]): row for row in thresholds}
    with plt.rc_context({"font.size": 9, "axes.titleweight": "bold"}):
        for paradigm in ("sharedreward", "trust", "ugr", "socialdoors"):
            group = [row for row in rows if row["paradigm"] == paradigm]
            fig, axes = plt.subplots(2, 2, figsize=(10, 7), constrained_layout=True)
            for axis, metric in zip(axes.flat, METRIC_FLAGS, strict=True):
                values = [
                    float(row[metric]) for row in group if row.get(metric) is not None
                ]
                threshold = lookup[(paradigm, metric)]
                if values:
                    bins: int | Any = min(30, max(5, int(math.sqrt(len(values))) + 1))
                    if metric == "tedana_rejected_components":
                        low, high = int(min(values)), int(max(values))
                        bins = np.arange(low - 0.5, high + 1.5, 1.0)
                    axis.hist(values, bins=bins, color="#4C78A8", edgecolor="white")
                    fence_name = (
                        "lower_fence"
                        if threshold["outlier_direction"] == "lower"
                        else "upper_fence"
                    )
                    fence = float(threshold[fence_name])
                    axis.axvline(fence, color="#C23B22", linewidth=2)
                    axis.text(
                        0.98,
                        0.95,
                        f"fence = {fence:.3g}\nn = {len(values)}",
                        transform=axis.transAxes,
                        ha="right",
                        va="top",
                    )
                else:
                    axis.text(0.5, 0.5, "No valid values", ha="center", va="center")
                axis.set_title(labels[metric])
                axis.set_ylabel("Runs")
                axis.set_xlabel(labels[metric])
            subtitle = (
                "task-socialdoors + task-doors (pooled)"
                if paradigm == "socialdoors"
                else f"task-{paradigm}"
            )
            fig.suptitle(f"RF1-SRA {paradigm} imaging QC\n{subtitle}", fontsize=13)
            fig.savefig(
                directory / f"{paradigm}_histograms.png",
                dpi=180,
                metadata={"Software": "rf1-sra-linux2 build_run_qc.py"},
            )
            plt.close(fig)


def generated_by(root: Path, project_root: Path) -> dict[str, Any]:
    path = root / "dataset_description.json"
    if not path.is_file():
        return {"metadata": "unavailable"}
    try:
        data = json.loads(path.read_text())
    except Exception:
        return {"metadata": relative_path(path, project_root), "error": "invalid JSON"}
    return {
        "metadata": relative_path(path, project_root),
        "name": data.get("Name", ""),
        "generated_by": data.get("GeneratedBy", []),
    }


def package_versions() -> dict[str, str]:
    versions: dict[str, str] = {}
    for name in ("numpy", "pandas", "nibabel", "scipy", "matplotlib", "openpyxl"):
        try:
            versions[name] = importlib.metadata.version(name)
        except importlib.metadata.PackageNotFoundError:
            versions[name] = "unavailable"
    return versions


def discover_template_mask(explicit: Path | None) -> Path:
    if explicit is not None:
        path = explicit.expanduser().resolve()
        if not path.is_file():
            raise ValueError(f"TemplateFlow brain mask not found: {path}")
        return path
    roots = [
        Path(os.environ.get("TEMPLATEFLOW_HOME", "/nonexistent")),
        Path("/ZPOOL/data/tools/templateflow"),
    ]
    candidates: set[Path] = set()
    name = "tpl-MNI152NLin6Asym_res-02_desc-brain_mask.nii.gz"
    for root in roots:
        if root.is_dir():
            candidates.update(root.rglob(name))
    if len(candidates) != 1:
        preview = ", ".join(str(path) for path in sorted(candidates)) or "none"
        raise ValueError(
            "Could not identify exactly one TemplateFlow MNI152NLin6Asym res-02 "
            f"brain mask ({preview}); pass --template-brain-mask explicitly"
        )
    return next(iter(candidates))


def print_summary(rows: list[dict[str, Any]], thresholds: list[dict[str, Any]]) -> None:
    print(f"Run inventory: {len(rows)}")
    print(f"Complete QC rows: {sum(row['qc_complete'] for row in rows)}")
    print(f"Incomplete QC rows: {sum(not row['qc_complete'] for row in rows)}")
    print(f"Any imaging-QC outlier: {sum(row['imaging_qc_outlier'] for row in rows)}")
    for task in sorted({row["task"] for row in rows}):
        group = [row for row in rows if row["task"] == task]
        print(
            f"task-{task}: total={len(group)} complete={sum(row['qc_complete'] for row in group)} "
            f"incomplete={sum(not row['qc_complete'] for row in group)} "
            f"tsnr={sum(row.get('tsnr_outlier') is True for row in group)} "
            f"fd={sum(row.get('fd_mean_outlier') is True for row in group)} "
            f"tedana={sum(row.get('tedana_outlier') is True for row in group)} "
            f"coverage={sum(row.get('brain_coverage_outlier') is True for row in group)} "
            f"any={sum(row['imaging_qc_outlier'] for row in group)}"
        )
    missing_counts: dict[str, int] = {}
    overlap_counts: dict[str, int] = {}
    for row in rows:
        for issue in filter(None, str(row["missing_metrics"]).split(";")):
            missing_counts[issue] = missing_counts.get(issue, 0) + 1
        flags = [
            metric for metric, flag in METRIC_FLAGS.items() if row.get(flag) is True
        ]
        if flags:
            combination = "+".join(flags)
            overlap_counts[combination] = overlap_counts.get(combination, 0) + 1
    if missing_counts:
        print("Missing/ambiguous inputs:")
        for issue, count in sorted(missing_counts.items()):
            print(f"  {issue}: {count}")
    if overlap_counts:
        print("Outlier-criterion overlap:")
        for combination, count in sorted(overlap_counts.items()):
            print(f"  {combination}: {count}")
    print("Thresholds:")
    for row in thresholds:
        fence = row[
            "lower_fence" if row["outlier_direction"] == "lower" else "upper_fence"
        ]
        shown = "NA" if fence is None else f"{float(fence):.10g}"
        print(
            f"  {row['paradigm']} {row['metric']}: n={row['n']} "
            f"{row['outlier_direction']}_fence={shown} outliers={row['n_outliers']}"
        )


def preflight_outputs(output_dir: Path, overwrite: bool) -> None:
    existing = [
        output_dir / relative
        for relative in CANONICAL_OUTPUTS
        if (output_dir / relative).exists()
    ]
    if existing and not overwrite:
        preview = "\n  ".join(str(path) for path in existing)
        raise ValueError(
            "Canonical QC outputs already exist; review them and rerun with --overwrite:\n  "
            + preview
        )


def commit_outputs(stage: Path, output_dir: Path) -> None:
    for relative in CANONICAL_OUTPUTS:
        source = stage / relative
        destination = output_dir / relative
        if not source.is_file():
            raise ValueError(f"staged QC output missing: {source}")
        destination.parent.mkdir(parents=True, exist_ok=True)
        os.replace(source, destination)


def run_build(args: argparse.Namespace) -> int:
    project_root = args.project_root.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    policy_path = args.policy.expanduser().resolve()
    exclusion_mask = args.exclusion_mask.expanduser().resolve()
    excluded_root = args.excluded_source_root.expanduser().resolve()
    policy = load_policy(policy_path)
    expected_exclusion_sha = policy["coverage"]["exclusion_mask_sha256"]
    if (
        not exclusion_mask.is_file()
        or sha256_file(exclusion_mask) != expected_exclusion_sha
    ):
        raise ValueError(
            "historical cerebellum/brainstem exclusion mask checksum mismatch"
        )
    template_mask = discover_template_mask(args.template_brain_mask)
    if not args.dry_run:
        preflight_outputs(output_dir, args.overwrite)
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix=".run-qc-stage-", dir=output_dir.parent
    ) as temp:
        stage = Path(temp)
        target_mask = stage / "reference" / TARGET_MASK_NAME
        coverage_provenance = build_target_mask(
            template_mask, exclusion_mask, target_mask
        )
        rows, inventory = build_rows(
            project_root,
            policy,
            target_mask,
            excluded_root,
            args.include_source_excluded,
        )
        if not rows:
            raise ValueError("no production BIDS echo-2 part-mag runs were discovered")
        thresholds = compute_thresholds(rows, policy)
        apply_thresholds(rows, thresholds, policy)
        pairs = build_socialdoors_pairs(rows)
        print_summary(rows, thresholds)
        if args.dry_run:
            print("DRY RUN: no canonical QC outputs were replaced.")
            return 0
        write_tsv(stage / "run_qc.tsv", rows, RUN_COLUMNS)
        write_tsv(stage / "thresholds.tsv", thresholds, THRESHOLD_COLUMNS)
        write_tsv(stage / "socialdoors_pair_qc.tsv", pairs, PAIR_COLUMNS)
        write_workbooks(stage / "spreadsheets", rows, thresholds, pairs)
        write_histograms(stage / "figures", rows, thresholds)
        provenance = {
            "schema_version": 1,
            "generated_at": utc_now(),
            "policy": "qc/qc_policy.json",
            "policy_sha256": sha256_file(policy_path),
            "inventory": inventory,
            "coverage": coverage_provenance,
            "derivatives": {
                "mriqc": generated_by(
                    project_root / "derivatives" / "mriqc", project_root
                ),
                "fmriprep": generated_by(
                    project_root / "derivatives" / "fmriprep", project_root
                ),
                "tedana": generated_by(
                    project_root / "derivatives" / "tedana", project_root
                ),
            },
            "python_packages": package_versions(),
            "outputs": [relative.as_posix() for relative in CANONICAL_OUTPUTS],
        }
        (stage / "provenance.json").write_text(
            json.dumps(provenance, indent=2, sort_keys=True) + "\n"
        )
        commit_outputs(stage, output_dir)
    print(f"QC outputs written under: {output_dir}")
    return 0


def parse_bool(value: str) -> bool | None:
    if value == "TRUE":
        return True
    if value == "FALSE":
        return False
    if value == "":
        return None
    raise ValueError(f"invalid boolean: {value}")


def canonical_native_rows(records: list[dict[str, str]]) -> list[dict[str, Any]]:
    integer_columns = {
        "tedana_total_components",
        "tedana_accepted_components",
        "tedana_rejected_components",
    }
    float_columns = {
        "tsnr",
        "fd_mean",
        "tedana_rejected_fraction",
        "brain_coverage_pct",
    }
    bool_columns = set(METRIC_FLAGS.values()) | {"imaging_qc_outlier", "qc_complete"}
    rows: list[dict[str, Any]] = []
    for record in records:
        row: dict[str, Any] = dict(record)
        for column in integer_columns:
            row[column] = int(record[column]) if record[column] else None
        for column in float_columns:
            row[column] = float(record[column]) if record[column] else None
        for column in bool_columns:
            row[column] = parse_bool(record[column])
        rows.append(row)
    return rows


def compare_number(actual: Any, expected: Any, atol: float = 1e-8) -> bool:
    if actual in (None, "") and expected in (None, ""):
        return True
    try:
        return math.isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=atol)
    except (TypeError, ValueError):
        return False


def workbook_run_keys(path: Path) -> set[tuple[str, str, str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("QC checking requires openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["runs"]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(iterator)]
    positions = [headers.index(name) for name in ("subject", "session", "task", "run")]
    result = {
        tuple(str(values[position]) for position in positions)
        for values in iterator
        if values[positions[0]] is not None
    }
    workbook.close()
    return result


def compare_live_rows(
    stored_rows: list[dict[str, Any]], live_rows: list[dict[str, Any]]
) -> list[str]:
    errors: list[str] = []
    key_columns = ("subject", "session", "task", "run")
    numeric_columns = (
        "tsnr",
        "fd_mean",
        "tedana_total_components",
        "tedana_accepted_components",
        "tedana_rejected_components",
        "tedana_rejected_fraction",
        "brain_coverage_pct",
    )
    text_columns = (
        "paradigm",
        "missing_metrics",
        "bids_bold",
        "mriqc_json",
        "tedana_metrics",
        "fmriprep_brain_mask",
    )
    stored_lookup = {
        tuple(row[column] for column in key_columns): row for row in stored_rows
    }
    live_lookup = {
        tuple(row[column] for column in key_columns): row for row in live_rows
    }
    if set(stored_lookup) != set(live_lookup):
        missing = sorted(set(live_lookup) - set(stored_lookup))
        extra = sorted(set(stored_lookup) - set(live_lookup))
        return [
            f"run inventory disagreement: missing={missing[:10]} extra={extra[:10]}"
        ]
    for key, live in live_lookup.items():
        stored = stored_lookup[key]
        for column in numeric_columns:
            if not compare_number(stored.get(column), live.get(column)):
                errors.append(f"upstream metric disagreement {key}: {column}")
        for column in text_columns:
            if stored.get(column, "") != live.get(column, ""):
                errors.append(f"upstream source disagreement {key}: {column}")
    return errors


def compare_pair_rows(
    stored_rows: list[dict[str, str]], expected_rows: list[dict[str, Any]]
) -> list[str]:
    errors: list[str] = []
    key_columns = ("subject", "session")
    stored_lookup = {
        tuple(row[column] for column in key_columns): row for row in stored_rows
    }
    expected_lookup = {
        tuple(row[column] for column in key_columns): row for row in expected_rows
    }
    if len(stored_lookup) != len(stored_rows):
        errors.append("socialdoors_pair_qc.tsv contains duplicate subject/session rows")
    if set(stored_lookup) != set(expected_lookup):
        errors.append("socialdoors_pair_qc.tsv key set disagrees with run_qc.tsv")
        return errors
    for key, expected in expected_lookup.items():
        stored = stored_lookup[key]
        for column in PAIR_COLUMNS:
            if stored.get(column, "") != str(output_value(expected.get(column))):
                errors.append(f"socialdoors pair disagreement {key}: {column}")
    return errors


def run_check(args: argparse.Namespace) -> int:
    project_root = args.project_root.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    policy = load_policy(args.policy.expanduser().resolve())
    errors: list[str] = []
    for relative in CANONICAL_OUTPUTS:
        if not (output_dir / relative).is_file():
            errors.append(f"missing canonical output: {relative}")
    if errors:
        for error in errors:
            print(f"CHECK FAILED: {error}")
        return 1
    records = read_tsv(output_dir / "run_qc.tsv")
    if not records or list(records[0]) != RUN_COLUMNS:
        errors.append("run_qc.tsv is empty or has a noncanonical header")
        rows: list[dict[str, Any]] = []
    else:
        rows = canonical_native_rows(records)
    keys = [(row["subject"], row["session"], row["task"], row["run"]) for row in rows]
    if len(keys) != len(set(keys)):
        errors.append("run_qc.tsv contains duplicate run keys")
    live_rows, _inventory = build_rows(
        project_root,
        policy,
        output_dir / "reference" / TARGET_MASK_NAME,
        args.excluded_source_root.expanduser().resolve(),
        args.include_source_excluded,
    )
    errors.extend(compare_live_rows(rows, live_rows))
    for row in rows:
        if row["paradigm"] != policy["task_map"].get(row["task"]):
            errors.append(f"incorrect paradigm mapping: {row['subject']} {row['task']}")
        if row["qc_complete"] and any(
            row.get(metric) is None for metric in METRIC_FLAGS
        ):
            errors.append(
                f"complete row lacks a primary metric: {tuple(row[name] for name in RUN_COLUMNS[:5])}"
            )
        if any(str(row[column]).startswith("/") for column in RUN_COLUMNS[-4:]):
            errors.append(
                f"absolute source path in tracked TSV: {row['subject']} {row['task']} {row['run']}"
            )
    expected_thresholds = compute_thresholds(rows, policy) if rows else []
    replay_rows = [dict(row) for row in rows]
    apply_thresholds(replay_rows, expected_thresholds, policy)
    stored_thresholds = read_tsv(output_dir / "thresholds.tsv")
    threshold_header_ok = (
        bool(stored_thresholds) and list(stored_thresholds[0]) == THRESHOLD_COLUMNS
    )
    if not threshold_header_ok:
        errors.append("thresholds.tsv is empty or has a noncanonical header")
    stored_lookup = (
        {(row["paradigm"], row["metric"]): row for row in stored_thresholds}
        if threshold_header_ok
        else {}
    )
    expected_lookup = {
        (row["paradigm"], row["metric"]): row for row in expected_thresholds
    }
    if set(stored_lookup) != set(expected_lookup):
        errors.append(
            "thresholds.tsv does not contain exactly four paradigms x four metrics"
        )
    else:
        for key, expected in expected_lookup.items():
            stored = stored_lookup[key]
            for column in (
                "n",
                "q1",
                "q3",
                "iqr",
                "lower_fence",
                "upper_fence",
                "n_outliers",
            ):
                if not compare_number(stored[column], expected[column]):
                    errors.append(f"threshold mismatch {key} {column}")
            for column in ("bids_tasks", "outlier_direction"):
                if stored[column] != str(expected[column]):
                    errors.append(f"threshold mismatch {key} {column}")
    for stored, expected in zip(rows, replay_rows, strict=True):
        for column in (
            *METRIC_FLAGS.values(),
            "imaging_qc_outlier",
            "outlier_reasons",
            "qc_complete",
            "qc_status",
        ):
            if stored.get(column) != expected.get(column):
                errors.append(
                    f"flag/status mismatch {stored['subject']} ses-{stored['session']} "
                    f"task-{stored['task']} run-{stored['run']}: {column}"
                )
    expected_pairs = build_socialdoors_pairs(rows)
    stored_pairs = read_tsv(output_dir / "socialdoors_pair_qc.tsv")
    pair_header_ok = bool(stored_pairs) and list(stored_pairs[0]) == PAIR_COLUMNS
    if not pair_header_ok:
        errors.append("socialdoors_pair_qc.tsv is empty or has a noncanonical header")
    else:
        errors.extend(compare_pair_rows(stored_pairs, expected_pairs))
    for paradigm in policy["paradigms"]:
        workbook = output_dir / "spreadsheets" / f"{paradigm}_qc.xlsx"
        expected_keys = {
            (row["subject"], row["session"], row["task"], row["run"])
            for row in rows
            if row["paradigm"] == paradigm
        }
        try:
            if workbook_run_keys(workbook) != expected_keys:
                errors.append(f"{workbook.name} run rows disagree with run_qc.tsv")
        except Exception as exc:  # noqa: BLE001 - report every malformed workbook.
            errors.append(f"{workbook.name} is structurally invalid: {exc}")
    try:
        provenance = json.loads((output_dir / "provenance.json").read_text())
        if provenance.get("policy_sha256") != sha256_file(
            args.policy.expanduser().resolve()
        ):
            errors.append("provenance policy checksum disagrees with qc_policy.json")
        target_sha = provenance.get("coverage", {}).get("target_mask_sha256")
        if target_sha != sha256_file(output_dir / "reference" / TARGET_MASK_NAME):
            errors.append("provenance target-mask checksum disagreement")
    except (OSError, ValueError, TypeError) as exc:
        errors.append(f"provenance.json is invalid: {exc}")
    incomplete = [row for row in rows if not row["qc_complete"]]
    print_summary(rows, expected_thresholds)
    if incomplete:
        print("Incomplete runs:")
        for row in incomplete:
            print(
                f"  sub-{row['subject']} ses-{row['session']} task-{row['task']} "
                f"run-{row['run']}: {row['missing_metrics']}"
            )
        errors.append(f"{len(incomplete)} run(s) have incomplete QC")
    if errors:
        for error in errors:
            print(f"CHECK FAILED: {error}")
        return 1
    print(
        f"CHECK PASSED: {len(rows)} acquired BIDS run(s) have complete, internally "
        "consistent canonical imaging QC outputs."
    )
    return 0


def build_parser() -> argparse.ArgumentParser:
    repo_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    def common(subparser: argparse.ArgumentParser) -> None:
        subparser.add_argument("--project-root", type=Path, default=repo_root)
        subparser.add_argument("--output-dir", type=Path, default=repo_root / "qc")
        subparser.add_argument(
            "--policy", type=Path, default=repo_root / "qc" / "qc_policy.json"
        )
        subparser.add_argument(
            "--excluded-source-root",
            type=Path,
            default=Path("/ZPOOL/data/sourcedata/sourcedata/rf1-sra-exclusions"),
        )
        subparser.add_argument("--include-source-excluded", action="store_true")

    build = subparsers.add_parser(
        "build", help="collect metrics and generate canonical QC outputs"
    )
    common(build)
    build.add_argument("--template-brain-mask", type=Path)
    build.add_argument(
        "--exclusion-mask",
        type=Path,
        default=repo_root
        / "qc"
        / "reference"
        / "source-cerebellum-brainstem_mask.nii.gz",
    )
    build.add_argument("--overwrite", action="store_true")
    build.add_argument("--dry-run", action="store_true")
    build.set_defaults(func=run_build)

    check = subparsers.add_parser(
        "check", help="verify canonical QC outputs and live run coverage"
    )
    common(check)
    check.set_defaults(func=run_check)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        return int(args.func(args))
    except (OSError, ValueError, RuntimeError) as exc:
        print(f"ERROR: {exc}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
